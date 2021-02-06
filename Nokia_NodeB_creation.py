import os
from sys import exit
from csv import DictReader
from openpyxl import load_workbook
from xml.etree.ElementTree import parse,Element,SubElement,Comment,tostring
from xml.dom import minidom
from datetime import datetime
from time import sleep
from subprocess import Popen
from colorama import init,Fore,Style

def prettify(raml):
    rough_string = tostring(raml, encoding='utf-8')
    reparsed = minidom.parseString(rough_string)
    return reparsed.toprettyxml(indent="     ")

def mo_creator(operation,version,distName,class_name,mo_params):
    if operation in ["create","update"]:
        mo_elem=Element("managedObject",version=version,distName=distName,operation=operation)
        mo_elem.set("class",class_name)
        for i in mo_params:
            if mo_params[i] !="":
                if i=="URAId":
                    uraid_list_mo=SubElement(mo_elem,"list",name="URAId")
                    uraid_list_p=SubElement(uraid_list_mo,"p").text=str(mo_params[i])
                elif i=="CControlPortID":
                    dnbap_list_mo=SubElement(mo_elem,"list",name="DNBAP")
                    dnbap_item_mo=SubElement(dnbap_list_mo,"item")
                    ccontrolportid_p=SubElement(dnbap_item_mo,"p",name=i).text=str(mo_params[i])
                else:
                    p_elem=SubElement(mo_elem,"p",name=i).text=str(mo_params[i])
            else:
                continue
            
        return mo_elem
    elif operation=="delete":
        mo_elem=Element("managedObject",version=version,distName=distName,operation=operation)
        mo_elem.set("class",class_name)
    else:
        print(Fore.RED+"Unknown operation")
    return mo_elem

def rows_handler(rows_list,mo_name=None):
    mo_elems_list=[]
    operations_supported=['create','delete','update']
    row_count=0
    for row in rows_list:
        operation=row['operation']
        mo_name=row['OBJECT']
        row_count+=1
        if operation in operations_supported:
            if mo_name in ["IPNB","WBTS"]:
                distName='PLMN-PLMN/'+'RNC-'+row['RNC']+'/'+mo_name+'-'+row[mo_name]
                del row['OBJECT'],row['RNC'],row[mo_name],row['operation']
            elif mo_name=="WCEL":
                distName='PLMN-PLMN/'+'RNC-'+row['RNC']+'/WBTS-'+row['WBTS']+'/WCEL-'+row[mo_name]
                del row['OBJECT'],row['RNC'],row['WBTS'],row[mo_name],row['operation']
            elif mo_name in ["ADJD","ADJE","ADJG","ADJI","ADJL","ADJS"]:
                distName='PLMN-PLMN/'+'RNC-'+row['RNC']+'/WBTS-'+row['WBTS']+'/WCEL-'+row['WCEL']+'/'+mo_name+'-'+row[mo_name]
                del row['OBJECT'],row['RNC'],row['WBTS'],row['WCEL'],row[mo_name],row['operation']
            else:
                print(Fore.RED+"OBJECT name in input file not supported. Please check input CSV file ROW number {}.".format(row_count))
                return False
        else:
            print(Fore.RED+"OPERATION name in input file not supported. Please check input CSV file ROW number {}.".format(row_count))
            return False
        if "" in row:
            del row['']
        mo_elem=mo_creator(operation=operation,version=rnc_sw_version,distName=distName,class_name=mo_name,mo_params=row)
        mo_elems_list.append(mo_elem)
    
    return mo_elems_list,mo_name

def excel_reader(wb):
    excel_mo_elems=[]
    mo_name_list=[]
    for ws in wb.worksheets:
        mo_name=ws.title
        mo_name_list.append(mo_name)
        all_rows=[]
        header=[cell.value for cell in ws[1]]
        row_count=0
        for each_row in ws.iter_rows(min_row=2):
            values={}
            for key,cell in zip(header,each_row):
                values[key]=str(cell.value)
            all_rows.append(values)
            row_count+=1
        mo_elems_list,mo_name=rows_handler(rows_list=all_rows,mo_name=mo_name)
        excel_mo_elems.append(Comment("Managed Objects for class {}".format(mo_name)))
        for mo_elem in mo_elems_list:
            excel_mo_elems.append(mo_elem)
    return excel_mo_elems,mo_name_list

def xml_creator(mo_elems_list):
    raml=Element("raml",version="2.0",xmlns="raml20.xsd")
    cmdata=SubElement(raml,"cmData",type="plan",name="3G Plan")
    header=SubElement(cmdata,"header")
    log=SubElement(header,"log",user="kazi.noor@nokia.com",dateTime=datetime.now().strftime("%c"),action="created",appInfo="Nokia RAN")

    for mo_elem in mo_elems_list:
        cmdata.append(mo_elem)
    return raml

def xml_save(raml_pretty,mo_name):
    time_now=datetime.now().strftime(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
    if not os.path.exists(os.path.join(os.getcwd(),'output')):
        os.mkdir(os.path.join(os.getcwd(),'output'))
    output_file_path=os.path.join(os.getcwd(),'output',mo_name+'_'+time_now+'.xml')
    with open(output_file_path,'w') as f_out:
        f_out.write(raml_pretty)
    print(Fore.GREEN+"New {} plan file created in output directory with file name -".format(mo_name))
    print(Fore.GREEN+Style.BRIGHT+output_file_path)
    Popen('explorer "{}"'.format(os.path.join(os.getcwd(),'output')))



def main():
    init(autoreset=True)
    print(Fore.CYAN+"Welcome to Nokia 3G Site Creation script ..")
    while True:
        print(Fore.YELLOW+"\nSupported Input File type -- :\n--> CSV       Input CSV file, each with specific MO type.\n--> DIR       Directory path containing CSV files for each MO\n--> Excel     Input EXCEL file, each sheet with specific MO type.\n\n--> Exit")
        input_text=input("Please enter file type for plan creation : ").strip().upper()
        
        if input_text=="CSV":
            input_file_path=input("Please enter the input CSV file path : ").strip('"')
            if not input_file_path.lower().endswith('csv'):
                print(Fore.RED+"Input file extension not recognized as a CSV file. Please try again.")
                continue
            f_in=open(input_file_path,'r',newline='')
            rows_dicts=DictReader(f_in)

            rows_handler_output=rows_handler(rows_list=rows_dicts)
            f_in.close()
            if rows_handler_output==False:
                continue            
            mo_elems_list,mo_name=rows_handler_output

            raml=xml_creator(mo_elems_list=mo_elems_list)


        elif input_text.startswith("DIR"):
            dirpath=input("Please enter the directory : ").strip('"').strip()
            dir_mo_elems=[]
            mo_name_list=[]

            for file in os.listdir(dirpath):
                if file.endswith('.csv'):
                    filepath=dirpath+'\\'+file
                    f_in=open(filepath,'r',newline='')
                    rows_dicts=DictReader(f_in)
                    rows_handler_output=rows_handler(rows_list=rows_dicts)
                    mo_elems_list,mo_name=rows_handler_output
                    dir_mo_elems.append(Comment("Managed Objects for class {}".format(mo_name)))
                    for mo_elem in mo_elems_list:
                        dir_mo_elems.append(mo_elem)
                    mo_name_list.append(mo_name)
                    f_in.close()
            
            raml=xml_creator(dir_mo_elems)
            mo_name='_'.join(mo_name_list)


        elif input_text=="EXCEL":
            input_file_path=input("Please enter the input EXCEL file path : ").strip('"')
            if not input_file_path.lower().endswith('xlsx'):
                print(Fore.RED+"Input file extension not recognized as a EXCEL file. Please try again.")
                continue
            wb=load_workbook(input_file_path)
            excel_mo_elems,mo_name_list=excel_reader(wb)
            raml=xml_creator(excel_mo_elems)
            mo_name='_'.join(mo_name_list)


        elif input_text=='EXIT':
            print(Fore.MAGENTA+"Thanks for using site creation script .. Goodbye!")
            sleep(1)
            exit()

        else:
            print(Fore.RED+"Wrong input. Please try again.")
            continue

        raml_pretty=prettify(raml)
        xml_save(raml_pretty=raml_pretty,mo_name=mo_name)
        print("\n")


try:
    rnc_sw_version='mcRNC18'
    main()
except Exception as e:
    print(Fore.RED+"Exception occured :")
    print(Fore.RED+str(e))
    r=input(Fore.RED+"Please press X to close.")











